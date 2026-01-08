import asyncio
import glob
import json
import os
import sys
from random import random
from time import sleep
from tkinter import filedialog

import numpy as np
import pandas
from openpyxl.reader.excel import load_workbook

import Constants
from Constants import Acties
from CreateInputConfig import modify_config, create_new_config, save_config
from CreatePicture import create_picture, create_picture_test
from LoadTerms import load_terms
from RulesConfig import get_all_colors_in_column
from Singleton import app
from openpyxl.utils import get_column_letter

from Utils import add_base_path


async def listActions():
    actions = [Acties.MAAK_NIEUWE_CONFIG]

    if app.get().input_config_file_name is not None:
        actions.append(Acties.VERANDER_CONFIG)

    actions.append(Acties.LAAD_BESTAANDE_CONFIG)

    if app.get().input_config_file_name is not None:
        actions.append(Acties.GENEREER_PLAATJES)

    actions.append(Acties.EXIT)

    app.get().ui.setList("Kies een actie: ", actions, callbackActies)
    app.get().ui.setActiveConfigName(app.get().input_config_file_name)
    await app.get().ui.refreshScreen()

async def callback_let_user_choose_config(config_name):
    app.get().input_config_file_name = config_name
    await listActions()

async def test_create_config_file():
    await app.get().ui.EmptyScreen()
    app.get().ui.setInput("Wat is de nieuwe naam van de configuratie?", test_create_config_file_name_found_cb)

    await app.get().ui.refreshScreen()

async def test_create_config_file_name_found_cb(config_name):
    cfg = {}
    if not config_name.lower().endswith(".json"):
        config_name += ".json"
    save_config(cfg, config_name)
    app.get().ui.setActiveConfigName(config_name)
    app.get().ui.setActiveConfig(cfg)
    await prompt_input_input_file_new()
    #input file name


async def config_state_machine():
    state_machine = app.get().ui.getConfigStateMachine()
    print(app.get().ui.activeConfig)
    if "input_file_name" not in state_machine:
        await prompt_input_input_file_new()
    elif "file_has_header" not in state_machine:
        await prompt_input_header_new()
    elif "column_name" not in state_machine:
        await prompt_input_termen_new()
    elif "width" not in state_machine:
        await prompt_input_width_new("Type de gewenste breedte van het gegenereerde plaatje in pixels")
    elif "height" not in state_machine:
        await prompt_input_height_new("Type de gewenste hoogte van het gegenereerde plaatje in pixels")
    elif "background_color" not in state_machine:
        await prompt_input_background_color("Type de gewenste kleur van de achtergrond van het plaatje als hexstring (bijv: \"#673489\"")
    elif "font" not in state_machine:
        await prompt_input_font()
    elif "font_size" not in state_machine:
        await prompt_input_font_size()
    elif "margin" not in state_machine:
        await prompt_input_margin()
    elif "types" not in state_machine:
        await prompt_input_types()
    else:
        df = pandas.read_excel(app.get().ui.activeConfig["input_file_name"])
        col_idx = df.columns.get_loc(app.get().ui.activeConfig["column_name"]) + 1  # openpyxl is 1-based
        app.get().ui.activeConfig["column_letter"] = get_column_letter(col_idx)

        save_config(app.get().ui.activeConfig, app.get().ui.activeConfigName)
        await listActions()


async def config_type_state_machine():
    state_machine = app.get().ui.getTypeStateMachine()
    if "name" not in state_machine:
        await prompt_input_type_name()
    elif "generated_image_text_color" not in state_machine:
        await prompt_input_type_text_color()
    elif "method" not in state_machine:
        await prompt_input_type_method()
    elif "column" not in state_machine:
        await prompt_input_type_column()
    elif "excel_file_cell_color" not in state_machine:
        await prompt_input_type_cell_color()
    else:
        app.get().ui.setConfigType()
        print(app.get().ui.activeConfig)
        await config_state_machine()

async def prompt_input_type_column():
    await app.get().ui.EmptyScreen()
    config = app.get().ui.getActiveConfig()
    df = pandas.read_excel(config["input_file_name"])
    app.get().ui.setList("Welke kolom bevat de kleur?", df.columns, prompt_input_type_column_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_type_cell_color():
    await app.get().ui.EmptyScreen()
    config = app.get().ui.getActiveConfig()
    configType = app.get().ui.getNewConfigType()
    wb = load_workbook(config["input_file_name"])
    ws = wb.active

    column = configType["column"]  # column to scan
    all_colors = get_all_colors_in_column(ws, column, wb, config["input_file_name"])
    all_colors_string = convert_excel_colors_to_string(all_colors)
    app.get().ui.setList("Welke kleur hoort bij dit type?", all_colors_string, prompt_input_type_cell_color_cb)
    await app.get().ui.refreshScreen()

def convert_excel_colors_to_string(list_of_color_tuples):
    return [f"{value}. {ctype}" for value, ctype in list_of_color_tuples]

async def prompt_input_type_method():
    await app.get().ui.EmptyScreen()
    app.get().ui.setList("Wil je bepalen via celkleur of celinhoud, of iets bij dit type hoort?", ["celkleur", "celinhoud"], prompt_input_type_method_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_type_text_color():
    await app.get().ui.EmptyScreen()
    app.get().ui.setInput("Wat moet de tekstkleur zijn van dit type? Type de gewenste kleur als hexstring (bijv: \"#673489\")", prompt_input_type_text_color_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_type_name():
    await app.get().ui.EmptyScreen()
    app.get().ui.setInput("Wat is de naam van dit type?", prompt_input_type_name_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_type_column_cb(user_input):
    config = app.get().ui.getActiveConfig()
    df = pandas.read_excel(config["input_file_name"])
    columnIndex = int(np.where(df.columns.values == user_input)[0][0])
    app.get().ui.setConfigTypeField("column", columnIndex)
    await config_type_state_machine()

async def prompt_input_type_cell_color_cb(user_input):
    config = app.get().ui.getActiveConfig()
    configType = app.get().ui.getNewConfigType()
    wb = load_workbook(config["input_file_name"])
    ws = wb.active

    column = configType["column"]  # column to scan
    all_colors = get_all_colors_in_column(ws, column, wb, config["input_file_name"])
    all_colors_string = convert_excel_colors_to_string(all_colors)
    idx = all_colors_string.index(user_input)

    app.get().ui.setConfigTypeField("excel_file_cell_color", all_colors[idx][1])
    app.get().ui.setConfigTypeField("excel_file_color_type", all_colors[idx][0])
    await config_type_state_machine()

async def prompt_input_type_method_cb(user_input):
    app.get().ui.setConfigTypeField("method", user_input)
    await config_type_state_machine()

async def prompt_input_type_name_cb(user_input):
    app.get().ui.setConfigTypeField("name", user_input)
    await config_type_state_machine()

async def prompt_input_type_text_color_cb(user_input):
    app.get().ui.setConfigTypeField("generated_image_text_color", user_input)
    await config_type_state_machine()


async def prompt_input_types():
    await app.get().ui.EmptyScreen()
    list_options = ["Nieuw type toevoegen", "Doorgaan naar volgende stap"]
    app.get().ui.setList("Een type zorgt ervoor dat bepaalde groepen, bepaalde tekstkleuren krijgen", list_options, prompt_input_types_cb)
    print(app.get().ui.activeConfig)
    await app.get().ui.refreshScreen()

async def prompt_input_types_cb(user_input):
    if user_input == "Nieuw type toevoegen":
        await prompt_input_type_name()
    elif user_input == "Doorgaan naar volgende stap":
        app.get().ui.setActiveConfigField("types")
        await config_state_machine()

async def prompt_input_margin():
    await app.get().ui.EmptyScreen()
    app.get().ui.setNumberInput("Type de gewenste marge in pixels", prompt_input_margin_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_font_size():
    await app.get().ui.EmptyScreen()
    app.get().ui.setNumberInput("Type de gewenste tekstgrootte", prompt_input_font_size_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_font():
    await app.get().ui.EmptyScreen()
    app.get().ui.setInput("Type de naam van het font dat je wil gebruiken.", prompt_input_font_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_background_color(prompt):
    await app.get().ui.EmptyScreen()
    app.get().ui.setInput(prompt, prompt_input_background_color_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_width_new(prompt):
    await app.get().ui.EmptyScreen()
    app.get().ui.setNumberInput(prompt, prompt_input_width_new_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_height_new(prompt):
    await app.get().ui.EmptyScreen()
    app.get().ui.setNumberInput(prompt, prompt_input_height_new_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_margin_cb(user_input):
    app.get().ui.setActiveConfigField("margin", int(user_input))
    await config_state_machine()

async def prompt_input_font_size_cb(user_input):
    app.get().ui.setActiveConfigField("font_size", int(user_input))
    await config_state_machine()

async def prompt_input_font_cb(user_input):
    app.get().ui.setActiveConfigField("font", user_input)
    await config_state_machine()

async def prompt_input_background_color_cb(user_input):
    app.get().ui.setActiveConfigField("background_color", user_input)
    await config_state_machine()

async def prompt_input_width_new_cb(user_input):
    app.get().ui.setActiveConfigField("width", int(user_input))
    await config_state_machine()

async def prompt_input_height_new_cb(user_input):
    app.get().ui.setActiveConfigField("height", int(user_input))
    await config_state_machine()

async def prompt_input_termen_new():
    cfg = app.get().ui.getActiveConfig()
    df = pandas.read_excel(cfg["input_file_name"])

    await app.get().ui.EmptyScreen()
    app.get().ui.setList("Selecteer de kolom die de termen bevat. Het opgegeven bestand bevat de volgende kolommen:", df.columns, prompt_input_termen_new_cb)
    await app.get().ui.refreshScreen()


async def prompt_input_termen_new_cb(user_input):
    app.get().ui.setActiveConfigField("column_name", user_input)
    await config_state_machine()


async def prompt_input_input_file_new(default = None):
    await app.get().ui.EmptyScreen()
    app.get().ui.setButtonInput("Kies het excel bestand om de plaatjes te genereren.", "Klik hier om de verkenner te openen", prompt_input_file_new_cb)
    #app.get().ui.setInput("Kies het excel bestand om de plaatjes te genereren. Druk op Enter om de verkenner te openen", prompt_input_file_new_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_file_new_cb():
    # Open file picker dialog
    file_path = filedialog.askopenfilename(
        title="Kies een bestand",
        filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*"))
    )
    app.get().ui.setActiveConfigField("input_file_name", file_path)
    await config_state_machine()

async def prompt_input_header_new():
    await app.get().ui.EmptyScreen()
    app.get().ui.setList("Is de eerste rij van het bestand de koptekst/header?", ["ja", "nee"], prompt_input_header_new_cb)
    await app.get().ui.refreshScreen()

async def prompt_input_header_new_cb(user_input):
    app.get().ui.setActiveConfigField("file_has_header", user_input)
    #app.get().ui.setConfigStateMachine("file_has_header")
    await config_state_machine()


async def let_user_choose_config():
    config_folder = add_base_path(Constants.CONFIG_FOLDER)
    cfg_files = glob.glob(os.path.join(config_folder, "*.json"))

    if len(cfg_files) == 0:
        app.get().ui.setMessage("Er bestaan nog geen configuratie bestanden.")
        await listActions()

    # for i, file_name in enumerate(cfg_files, start=1):
    #     print(f"{i}. {print_config(file_name)}", "")

    app.get().ui.setList("Kies een input configuratie bestand:", cfg_files, callback_let_user_choose_config)
    await app.get().ui.refreshScreen()


async def slowTask(app, termen, file_name, percent):
    slowTask.total = len(termen)
    create_picture(termen[0], file_name, app)
    await app.increment(percent)
    await asyncio.sleep(0.01) #give time to UI thread to update.
    if len(termen) > 1:
        app.run_worker(slowTask(app, termen[1:], file_name, percent))
    else:
        app.setShowProgressBar(False)
        await listActions()

async def callbackActies(actie):
    match actie:
        case Acties.EXIT:
            await app.get().ui.action_quit()
        case Acties.LAAD_BESTAANDE_CONFIG:
            await let_user_choose_config()
        case Acties.VERANDER_CONFIG:
            modify_config(callbackActies.input_config_file_name)
        case Acties.MAAK_NIEUWE_CONFIG:
            await test_create_config_file()
            #input_config_file_name = create_new_config()
        case Acties.GENEREER_PLAATJES:
            with open(app.get().input_config_file_name) as f:
                cfg = json.load(f)
            termen = load_terms(cfg)
            #for term in termen:
                #await create_picture(term, cfg, app.get())
                #asyncio.create_task(create_picture(term, cfg, app.get()))

            await app.get().ui.EmptyScreen()
            app.get().ui.setShowProgressBar(True)
            await app.get().ui.refreshScreen()
            percent = 1/len(termen) * 100
            asyncio.create_task(slowTask(app.get().ui, termen, cfg, percent))
            #for term in termen:
                #sleep(2)
                #await app.get().ui.increment()
                #app.get().ui.run_worker(create_picture(term, cfg, app.get().ui), process=True)
                #asyncio.create_task(create_picture(term, cfg, app.get().ui))
                #asyncio.create_task(slowTask(app.get().ui))

