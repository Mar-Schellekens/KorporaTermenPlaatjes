import pandas
from openpyxl import load_workbook
from Constants import CfgFields, TypesMethod, StateMachines
from Model import Model
from Term import Term
from View import View


def classify_by_color(typ, cell, column_index):
    if typ[CfgFields.TYPES_EXCEL_FILE_COLOR_TYPE.value] == "theme":
        if str(cell[column_index].fill.start_color.theme) == typ[CfgFields.TYPES_EXCEL_FILE_CELL_COLOR.value]:
            return True
    elif typ[CfgFields.TYPES_EXCEL_FILE_COLOR_TYPE.value] == "rgb":
        if typ[CfgFields.TYPES_EXCEL_FILE_CELL_COLOR.value] == "geen kleur":
            if cell[column_index].fill.patternType is None:
                return True
        elif cell[column_index].fill.start_color.rgb == typ[CfgFields.TYPES_EXCEL_FILE_CELL_COLOR.value]:
            return True
    return False

def classify_by_value(typ, cell, column_index):
    if cell[column_index].value == typ[CfgFields.TYPES_MATCH_STRING.value]:
        return True
    return False

def classify_cell(cell, types):
    for typ in types:
        column_index = typ[CfgFields.TYPES_COLUMN.value]
        if typ[CfgFields.TYPES_METHOD.value] == "celkleur":
            if classify_by_color(typ, cell, column_index):
                return typ[CfgFields.TYPES_GENERATED_IMAGE_TEXT_COLOR.value]
        elif typ[CfgFields.TYPES_METHOD.value] == TypesMethod.CEL_INHOUD:
            if classify_by_value(typ, cell, column_index):
                return typ[CfgFields.TYPES_GENERATED_IMAGE_TEXT_COLOR.value]

    return "#000000"



def load_terms(cfg):
    try:
        df = pandas.read_excel(cfg[CfgFields.INPUT_FILE_NAME.value])
    except FileNotFoundError as e:
        View.get().set_error_message(str(e))
        Model.current_state_machine = StateMachines.MAIN_MENU
        return None

    termen = df[cfg[CfgFields.COLUMN_NAME.value]].values
    wb = load_workbook(cfg[CfgFields.INPUT_FILE_NAME.value])
    ws = wb.active
    colors = [classify_cell(row, cfg[CfgFields.TYPES.value]) for row in ws.iter_rows()]

    if cfg[CfgFields.FILE_HAS_HEADER.value]:
        colors = colors[1:] #Remove header row

    termObjects = []
    for term, color in zip(termen, colors):
        termObjects.append(Term(term, color))

    return termObjects