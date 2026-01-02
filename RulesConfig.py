from typing import Any

from colorist import hex
from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import column_index_from_string

def let_user_choose_ordered_list(list):
    for i, item in enumerate(list, start=1):
        print(f"{i}. {str(item)}", "")

    while True:
        choice = input("\nType het nummer van uw keuze: ").strip()
        if choice.isdigit():
            index = int(choice) - 1
            if 0 <= index < len(list):
                item = list[index]
                break
        print("Invalide keuze. Probeer opnieuw.")

    return item


def get_cell_color(cell, workbook):
    """Returns the RGB hex color of a cell, handling RGB, theme, and indexed colors."""
    color = cell.fill.start_color
    if color is None:
        return None

    if color.type == "rgb":
        return "rgb", color.rgb  # Already in hex
    elif color.type == "theme":
        return "theme", str(color.theme)
    return None

def get_all_colors_in_column(sheet, col_index, workbook, excel_file_name):
    """Returns a set of all unique colors in a column."""
    colors = []

    for row in sheet.iter_rows(min_col=col_index, max_col=col_index):
        cell = row[0]
        type, value = get_cell_color(cell, workbook)
        if (type, value) not in colors:
            colors.append((type, value))

    return colors

def prompt_input(prompt_text, default=None, cast=str, validator=None):
    """
    Prompt the user for input with optional default, type casting, and validation.
    """
    if default is not None:
        prompt_text = f"{prompt_text} [{default}]: "
    else:
        prompt_text = f"{prompt_text}: "

    while True:
        user_input = input(prompt_text).strip()
        if not user_input and default is not None:
            return default
        elif not user_input:
            return None  # allow empty input
        try:
            value = cast(user_input)
            if validator and not validator(value):
                print("Invalid value. Please try again.")
                continue
            return value
        except ValueError:
            print(f"Invalid input type. Expected {cast.__name__}.")


def is_valid_hex_color(s):
    """Validate hex color like #FFF or #FF8040"""
    s = s.lstrip("#")
    return len(s) in (3, 6) and all(c in "0123456789abcdefABCDEF" for c in s)

def prompt_input_cell_color(excel_file_name, column_letter):
    print("De volgende kleuren zijn aanwezig in die kolom:")

    wb = load_workbook(excel_file_name)
    ws = wb.active

    column = column_letter  # column to scan
    all_colors = get_all_colors_in_column(ws, column, wb, excel_file_name)

    color = let_user_choose_ordered_list(all_colors)
    return color

def edit_rules(rules, excel_file_name):
    final_rules = []

    # Edit existing defaults
    if rules is not None:
        for r in rules:
            print(f"\n of iets type '{r['type']}' is wordt bepaald door de volgende methode: ")
            print(f"bepaald door: {r['method']}")
            print(f"In kolom: {r['column']}")
            print(f"gelijk aan kleur: {r['color']}")

            action = input(
                "Type \"wijzig\" om te wijzigen, druk op \"Enter\" om te behouden, of type \"verwijder\" om het te verwijderen. ").strip().lower()
            if action == "verwijder":
                print(f"Heb '{r['type']}' Verwijderd")
                continue
            if action == "wijzig":
                type = r["type"]
                method = prompt_input("Wil je bepalen via \"celkleur\" of \"celinhoud\"? (celinhoud wordt nog niet ondersteund)", default=r.get("method"))
                column = prompt_input("Welke kolom?", default=r.get("column"))
                color_type, color = prompt_input_cell_color(excel_file_name, column)

                name = prompt_input("Type nieuwe naam of druk op \"Enter\" om huidige te behouden ", default=r['name'])
                color = prompt_input(
                    "Type nieuwe kleur als hex (bijv. #FF8040), of druk op \"Enter\" om huidige te behouden.",
                    default=r['color'], validator=is_valid_hex_color)
                final_rules.append({"name": name, "color": color})
            if action == "":
                final_rules.append(r)

    # Add new types
    while True:
        print("\nAdd a new type (leave name blank to finish):")
        name = prompt_input("Enter name")
        if not name:
            break
        color = prompt_input("Enter color (hex, e.g., #FF8040)", validator=is_valid_hex_color)
        if not color:
            print("Color is required for a new type.")
            continue
        final_rules.append({"name": name, "color": color})

    return final_rules


if __name__ == "__main__":
    main()