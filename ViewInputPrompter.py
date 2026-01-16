#This is an interface between control and view, part of the view layer, and should never import ControlInputHandler to prevent circular import.

import glob
import os
import pandas
from openpyxl.reader.excel import load_workbook
import Constants
import Utils
from Constants import TypesMethod, CfgFields, TYPE_FIELDS
from Model import Model
from RulesConfig import get_all_colors_in_column
from View import View
from Utils import add_base_path, convert_excel_colors_to_string

class ViewInputPrompter:
    async def exit(self):
        await View.get().action_quit()

    async def show_modify_config(self, cb):
        await View.get().empty_screen()
        options = [f.display for f in CfgFields if f.display is not None and f not in TYPE_FIELDS]
        View.get().set_list("Welk deel van de configuratie wil je aanpassen?", options, cb)
        await View.get().refresh_screen()

    async def show_main_menu(self, actions, cb):
        await View.get().empty_screen()
        View.get().set_list("Kies een actie: ", actions, cb)
        await View.get().refresh_screen()

    async def show_progress_bar(self, cb):
        await View.get().empty_screen()
        View.get().set_show_progress_bar(True, cb)
        await View.get().refresh_screen()

    async def name_new_config(self, cb):
        await View.get().empty_screen()
        View.get().set_text_input("Wat is de nieuwe naam van de configuratie?", cb)
        await View.get().refresh_screen()

    async def type_column(self, cb):
        await View.get().empty_screen()
        config = Model.get().get_active_cfg()
        df = pandas.read_excel(config["input_file_name"])

        if Model.get().new_config_type[Constants.CfgFields.TYPES_METHOD.value] == TypesMethod.CEL_KLEUR:
            prompt = "Welke kolom bevat de kleur die bepaalt of iets bij dit type hoort?"
        elif Model.get().new_config_type[Constants.CfgFields.TYPES_METHOD.value] == TypesMethod.CEL_INHOUD:
            prompt = "Welke kolom bevat de tekst die bepaalt of iets bij dit type hoort?"
        else:
            raise Exception("Config type method field somehow has an unvalid value")

        View.get().set_list(prompt, df.columns, cb)
        await View.get().refresh_screen()

    async def cell_content_type(self, cb):
        await View.get().empty_screen()
        config = Model.get().get_active_cfg()
        configType = Model.get().get_new_config_type()
        wb = load_workbook(config["input_file_name"])
        ws = wb.active
        column = configType["column"]  # column to scan

        values = []
        for i, row in enumerate(ws):
            if Model.get().active_config[Constants.CfgFields.FILE_HAS_HEADER.value] == "ja":
                if i == 0:
                    continue
            value = row[column].value
            if value not in values and value is not None:
                values.append(value)

        if len(values) == 0:
            raise Exception("Er zijn geen waardes aanwezig in de aangegeven kolom")
        View.get().set_list("Welke waarde moet een cel bevatten om bij dit type te horen?", values, cb)
        await View.get().refresh_screen()


    async def cell_color_type(self, cb):
        await View.get().empty_screen()
        config = Model.get().get_active_cfg()
        configType = Model.get().get_new_config_type()
        wb = load_workbook(config["input_file_name"])
        ws = wb.active

        column = configType["column"]  # column to scan
        all_colors = get_all_colors_in_column(ws, column, wb)

        all_colors_string, hex_colors = convert_excel_colors_to_string(all_colors)
        View.get().set_list("Welke kleur hoort bij dit type?", all_colors_string, cb, colors=hex_colors)
        await View.get().refresh_screen()

    async def type_method(self, cb):
        await View.get().empty_screen()
        View.get().set_list("Wil je bepalen via celkleur of celinhoud, of iets bij dit type hoort?", [TypesMethod.CEL_KLEUR, TypesMethod.CEL_INHOUD], cb)
        await View.get().refresh_screen()

    async def type_text_color(self, cb):
        await View.get().empty_screen()
        View.get().set_text_input("Wat moet de tekstkleur zijn van dit type? Type de gewenste kleur als hexcode (bijv: #673489). Zie www.htmlcolorcodes.com om de kleur code te krijgen.", cb)
        await View.get().refresh_screen()

    async def type_name(self, cb):
        await View.get().empty_screen()
        View.get().set_text_input("Wat is de naam van dit type?", cb)
        await View.get().refresh_screen()

    async def types(self, cb):
        await View.get().empty_screen()

        type_names = []
        if Constants.CfgFields.TYPES.value in Model.get().active_config:
            for typ in Model.get().active_config[Constants.CfgFields.TYPES.value]:
                type_names.append(typ[Constants.CfgFields.TYPES_NAME.value])


        options = [Constants.TypesMenu.ADD, Constants.TypesMenu.CONT]

        types = []
        if Constants.CfgFields.TYPES.value in Model.get().active_config:
            types = Model.get().active_config[Constants.CfgFields.TYPES.value]

        if len(types)>0:
            options.append(Constants.TypesMenu.DEL)

        View.get().set_type_overview(type_names, "Een type zorgt ervoor dat bepaalde groepen, bepaalde tekstkleuren krijgen", list(
            options), cb)
        await View.get().refresh_screen()

    async def margin(self, cb):
        await View.get().empty_screen()
        View.get().set_number_input("Type de gewenste marge in pixels", cb)
        await View.get().refresh_screen()

    async def font_size(self, cb):
        await View.get().empty_screen()
        View.get().set_number_input("Type de gewenste tekstgrootte", cb)
        await View.get().refresh_screen()

    async def font(self, cb):
        await View.get().empty_screen()
        View.get().set_text_input("Type de naam van het font dat je wil gebruiken.", cb)
        await View.get().refresh_screen()

    async def background_color(self, cb):
        await View.get().empty_screen()
        View.get().set_text_input("Type de gewenste kleur van de achtergrond van het plaatje als hexcode (bijv: #673489). Zie www.htmlcolorcodes.com om de kleur code te krijgen.", cb)
        await View.get().refresh_screen()

    async def width(self, cb):
        await View.get().empty_screen()
        View.get().set_number_input("Type de gewenste breedte van het gegenereerde plaatje in pixels", cb)
        await View.get().refresh_screen()

    async def height(self, cb):
        await View.get().empty_screen()
        View.get().set_number_input("Type de gewenste hoogte van het gegenereerde plaatje in pixels", cb)
        await View.get().refresh_screen()

    async def column_name(self, cb):
        cfg = Model.get().get_active_cfg()
        df = pandas.read_excel(cfg["input_file_name"])

        await View.get().empty_screen()
        View.get().set_list("Selecteer de kolom die de termen bevat. Het opgegeven bestand bevat de volgende kolommen:", df.columns, cb)
        await View.get().refresh_screen()

    async def input_excel_file(self, cb):
        await View.get().empty_screen()
        View.get().set_button_input("Kies het excel bestand om de plaatjes te genereren.", "Klik hier om de verkenner te openen", cb)
        await View.get().refresh_screen()

    async def header(self, cb):
        await View.get().empty_screen()
        View.get().set_list("Is de eerste rij van het bestand de koptekst/header?", ["ja", "nee"], cb)
        await View.get().refresh_screen()

    async def user_choose_config(self, cb, cb_no_configs):
        config_folder = add_base_path(Constants.CONFIG_FOLDER)
        cfg_files = glob.glob(os.path.join(config_folder, "*.json"))


        if len(cfg_files) == 0:
            View.get().set_message("Er bestaan nog geen configuratie bestanden.")
            await cb_no_configs()

        names = map(Utils.get_file_name_from_path, cfg_files)
        View.get().set_list("Kies een input configuratie bestand:", names, cb)
        await View.get().refresh_screen()

    async def delete_type(self, cb):
        await View.get().empty_screen()
        types_raw = Model.get().active_config[CfgFields.TYPES.value]
        types = [t[CfgFields.TYPES_NAME.value]  for t in types_raw]
        View.get().set_list("Welk type wil je verwijderen?", types, cb)
        await View.get().refresh_screen()







